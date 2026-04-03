from fastapi import FastAPI, Depends, HTTPException, status, Request, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import Optional, List
import os
import csv
import io
import json
import time
from urllib.parse import quote
from openpyxl import Workbook

from app.database import get_db, init_db
from app.models import Partner, Client, LaserSession, Tattoo, IntakeSubmission
from app.schemas import (
    PartnerCreate, PartnerUpdate, PartnerResponse,
    ClientCreate, ClientUpdate, ClientResponse, ClientListResponse,
    TattooCreate, TattooUpdate, TattooResponse, TattoosListResponse,
    LaserSessionCreate, LaserSessionUpdate, LaserSessionResponse, LaserSessionsListResponse,
    LoginRequest, LoginResponse, ClientStatus,
    PublicBookingCreate, PublicBookingResponse
)
from app.auth import (
    verify_password, create_session_token, get_current_user,
    require_auth, SESSION_COOKIE_NAME
)

# Initialize FastAPI app
app = FastAPI(title="TS Laser CRM", version="1.0.0")

# Setup static files and templates
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


@app.on_event("startup")
def startup_event():
    """Initialize database on startup"""
    init_db()


# ============== Auth Routes ==============

@app.post("/api/login", response_model=LoginResponse)
def login(request: LoginRequest):
    """Login with password"""
    if verify_password(request.password):
        response = JSONResponse(content={"success": True, "message": "РЈСЃРїРµС€РЅС‹Р№ РІС…РѕРґ"})
        token = create_session_token()
        response.set_cookie(
            key=SESSION_COOKIE_NAME,
            value=token,
            httponly=True,
            max_age=60 * 60 * 24 * 7,  # 7 days
            samesite="lax"
        )
        return response
    raise HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="РќРµРІРµСЂРЅС‹Р№ РїР°СЂРѕР»СЊ"
    )


@app.post("/api/logout")
def logout():
    """Logout - clear session"""
    response = JSONResponse(content={"success": True, "message": "Р’С‹С…РѕРґ РІС‹РїРѕР»РЅРµРЅ"})
    response.delete_cookie(key=SESSION_COOKIE_NAME)
    return response


@app.get("/api/auth/check")
def check_auth(request: Request):
    """Check if user is authenticated"""
    is_authenticated = get_current_user(request)
    return {"authenticated": is_authenticated}


# ============== Partner Routes ==============

@app.get("/api/partners", response_model=List[PartnerResponse])
def get_partners(
    request: Request,
    search: Optional[str] = None,
    type_filter: Optional[str] = None,
    sort_by: str = "name",
    sort_order: str = "asc",
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Get all partners with optional search, filter and sorting"""
    query = db.query(Partner)
    
    # Filter by type
    if type_filter:
        query = query.filter(Partner.type == type_filter)
    
    # Sorting
    sort_column = getattr(Partner, sort_by, Partner.name)
    if sort_order == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())
    
    results = query.all()
    
    # Search filter in Python (SQLite lower() doesn't work with Cyrillic)
    if search:
        search_lower = search.lower()
        results = [p for p in results if p.name and search_lower in p.name.lower()]
    
    return results


@app.post("/api/partners", response_model=PartnerResponse, status_code=status.HTTP_201_CREATED)
def create_partner(
    partner: PartnerCreate,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Create a new partner"""
    db_partner = Partner(**partner.model_dump())
    db.add(db_partner)
    db.commit()
    db.refresh(db_partner)
    return db_partner


@app.get("/api/partners/{partner_id}", response_model=PartnerResponse)
def get_partner(
    partner_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Get partner by ID"""
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="РџР°СЂС‚РЅРµСЂ РЅРµ РЅР°Р№РґРµРЅ")
    return partner


@app.put("/api/partners/{partner_id}", response_model=PartnerResponse)
def update_partner(
    partner_id: int,
    partner_update: PartnerUpdate,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Update partner"""
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="РџР°СЂС‚РЅРµСЂ РЅРµ РЅР°Р№РґРµРЅ")
    
    update_data = partner_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(partner, field, value)
    
    db.commit()
    db.refresh(partner)
    return partner


@app.delete("/api/partners/{partner_id}")
def delete_partner(
    partner_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Delete partner"""
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="РџР°СЂС‚РЅРµСЂ РЅРµ РЅР°Р№РґРµРЅ")
    
    db.delete(partner)
    db.commit()
    return {"success": True, "message": "РџР°СЂС‚РЅРµСЂ СѓРґР°Р»РµРЅ"}


# ============== Client Routes ==============

@app.get("/api/clients", response_model=List[ClientListResponse])
def get_clients(
    request: Request,
    search: Optional[str] = None,
    status_filter: Optional[str] = None,
    partner_filter: Optional[int] = None,
    sort_by: str = "name",
    sort_order: str = "asc",
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Get all clients with optional search, filter and sorting"""
    query = db.query(Client)
    
    # Filter by status (can be done in SQL)
    if status_filter:
        query = query.filter(Client.status == status_filter)
    
    # Filter by partner
    if partner_filter:
        query = query.filter(Client.referral_partner_id == partner_filter)
    
    # Sorting
    sort_column = getattr(Client, sort_by, Client.name)
    if sort_order == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())
    
    results = query.all()
    
    # Search filter in Python (SQLite lower() doesn't work with Cyrillic)
    if search:
        search_lower = search.lower()
        search_digits = ''.join(filter(str.isdigit, search))
        
        filtered = []
        for client in results:
            # Check name (case-insensitive)
            if client.name and search_lower in client.name.lower():
                filtered.append(client)
            # Check phone
            elif search_digits and client.phone and search_digits in client.phone:
                filtered.append(client)
        results = filtered
    
    return results


@app.post("/api/clients", response_model=ClientResponse, status_code=status.HTTP_201_CREATED)
def create_client(
    client: ClientCreate,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Create a new client"""
    client_data = client.model_dump()
    client_data['status'] = client_data['status'].value if hasattr(client_data['status'], 'value') else client_data['status']
    db_client = Client(**client_data)
    db.add(db_client)
    db.commit()
    db.refresh(db_client)
    
    # Get partner name if exists
    response_data = ClientResponse.model_validate(db_client)
    if db_client.referral_partner_id:
        partner = db.query(Partner).filter(Partner.id == db_client.referral_partner_id).first()
        if partner:
            response_data.referral_partner_name = partner.name
    
    return response_data


@app.get("/api/clients/{client_id}", response_model=ClientResponse)
def get_client(
    client_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Get client by ID"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    # Get partner name if exists
    response_data = ClientResponse.model_validate(client)
    if client.referral_partner_id:
        partner = db.query(Partner).filter(Partner.id == client.referral_partner_id).first()
        if partner:
            response_data.referral_partner_name = partner.name
    
    return response_data


@app.put("/api/clients/{client_id}", response_model=ClientResponse)
def update_client(
    client_id: int,
    client_update: ClientUpdate,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Update client"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    update_data = client_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field == 'status' and hasattr(value, 'value'):
            value = value.value
        setattr(client, field, value)
    
    db.commit()
    db.refresh(client)
    
    # Get partner name if exists
    response_data = ClientResponse.model_validate(client)
    if client.referral_partner_id:
        partner = db.query(Partner).filter(Partner.id == client.referral_partner_id).first()
        if partner:
            response_data.referral_partner_name = partner.name
    
    return response_data


@app.delete("/api/clients/{client_id}")
def delete_client(
    client_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Delete client"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    db.delete(client)
    db.commit()
    return {"success": True, "message": "РљР»РёРµРЅС‚ СѓРґР°Р»РµРЅ"}


# ============== Tattoo Routes ==============

@app.get("/api/clients/{client_id}/tattoos", response_model=TattoosListResponse)
def get_client_tattoos(
    client_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Get all tattoos for a client"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    tattoos = db.query(Tattoo).filter(
        Tattoo.client_id == client_id
    ).order_by(Tattoo.name.asc()).all()
    
    return TattoosListResponse(
        tattoos=[TattooResponse.model_validate(t) for t in tattoos],
        client_name=client.name,
        client_id=client.id
    )


@app.post("/api/clients/{client_id}/tattoos", response_model=TattooResponse, status_code=status.HTTP_201_CREATED)
def create_tattoo(
    client_id: int,
    tattoo: TattooCreate,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Create a new tattoo for a client"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    db_tattoo = Tattoo(
        client_id=client_id,
        **tattoo.model_dump()
    )
    db.add(db_tattoo)
    db.commit()
    db.refresh(db_tattoo)
    return db_tattoo


@app.get("/api/tattoos/{tattoo_id}", response_model=TattooResponse)
def get_tattoo(
    tattoo_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Get tattoo by ID"""
    tattoo = db.query(Tattoo).filter(Tattoo.id == tattoo_id).first()
    if not tattoo:
        raise HTTPException(status_code=404, detail="РўР°С‚СѓРёСЂРѕРІРєР° РЅРµ РЅР°Р№РґРµРЅР°")
    return tattoo


@app.put("/api/tattoos/{tattoo_id}", response_model=TattooResponse)
def update_tattoo(
    tattoo_id: int,
    tattoo_update: TattooUpdate,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Update tattoo"""
    tattoo = db.query(Tattoo).filter(Tattoo.id == tattoo_id).first()
    if not tattoo:
        raise HTTPException(status_code=404, detail="РўР°С‚СѓРёСЂРѕРІРєР° РЅРµ РЅР°Р№РґРµРЅР°")
    
    update_data = tattoo_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(tattoo, field, value)
    
    db.commit()
    db.refresh(tattoo)
    return tattoo


@app.delete("/api/tattoos/{tattoo_id}")
def delete_tattoo(
    tattoo_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Delete tattoo"""
    tattoo = db.query(Tattoo).filter(Tattoo.id == tattoo_id).first()
    if not tattoo:
        raise HTTPException(status_code=404, detail="РўР°С‚СѓРёСЂРѕРІРєР° РЅРµ РЅР°Р№РґРµРЅР°")
    
    db.delete(tattoo)
    db.commit()
    return {"success": True, "message": "РўР°С‚СѓРёСЂРѕРІРєР° СѓРґР°Р»РµРЅР°"}


# ============== LaserSession Routes ==============

@app.get("/api/clients/{client_id}/sessions", response_model=LaserSessionsListResponse)
def get_client_sessions(
    client_id: int,
    tattoo_filter: Optional[str] = None,
    sort_by: str = "session_date",
    sort_order: str = "asc",
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Get all laser sessions for a client"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    # Build query with sorting
    query = db.query(LaserSession).filter(LaserSession.client_id == client_id)
    
    # Filter by tattoo name
    if tattoo_filter:
        query = query.filter(LaserSession.tattoo_name == tattoo_filter)
    
    # Determine sort column
    sort_columns = {
        "tattoo_name": LaserSession.tattoo_name,
        "session_number": LaserSession.session_number,
        "session_date": LaserSession.session_date,
    }
    sort_column = sort_columns.get(sort_by, LaserSession.session_date)
    
    # Apply sort order
    if sort_order == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())
    
    sessions = query.all()
    
    # Calculate total flashes (for filtered results)
    total_flashes = sum(s.flashes_count or 0 for s in sessions)
    
    return LaserSessionsListResponse(
        sessions=[LaserSessionResponse.model_validate(s) for s in sessions],
        total_flashes=total_flashes,
        client_name=client.name,
        client_id=client.id
    )


@app.post("/api/clients/{client_id}/sessions", response_model=LaserSessionResponse, status_code=status.HTTP_201_CREATED)
def create_session(
    client_id: int,
    session: LaserSessionCreate,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Create a new laser session for a client"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    db_session = LaserSession(client_id=client_id, **session.model_dump())
    db.add(db_session)
    db.commit()
    db.refresh(db_session)
    return db_session


@app.get("/api/sessions/{session_id}", response_model=LaserSessionResponse)
def get_session(
    session_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Get laser session by ID"""
    session = db.query(LaserSession).filter(LaserSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="РЎРµР°РЅСЃ РЅРµ РЅР°Р№РґРµРЅ")
    return session


@app.put("/api/sessions/{session_id}", response_model=LaserSessionResponse)
def update_session(
    session_id: int,
    session_update: LaserSessionUpdate,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Update laser session"""
    session = db.query(LaserSession).filter(LaserSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="РЎРµР°РЅСЃ РЅРµ РЅР°Р№РґРµРЅ")
    
    update_data = session_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(session, field, value)
    
    db.commit()
    db.refresh(session)
    return session


@app.delete("/api/sessions/{session_id}")
def delete_session(
    session_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Delete laser session"""
    session = db.query(LaserSession).filter(LaserSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="РЎРµР°РЅСЃ РЅРµ РЅР°Р№РґРµРЅ")
    
    db.delete(session)
    db.commit()
    return {"success": True, "message": "РЎРµР°РЅСЃ СѓРґР°Р»РµРЅ"}


# ============== Export Routes ==============

def format_date(d):
    """Format date for export"""
    return d.strftime("%d.%m.%Y") if d else ""

def format_datetime(dt):
    """Format datetime for export"""
    return dt.strftime("%d.%m.%Y %H:%M") if dt else ""

def get_status_label(status_value):
    """Get status label in Russian"""
    status_map = {
        'active': 'Р’ РїСЂРѕС†РµСЃСЃРµ',
        'in_progress': 'Р’ РїСЂРѕС†РµСЃСЃРµ',
        'completed': 'Р—Р°РІРµСЂС€РµРЅРѕ',
        'stopped': 'РџРµСЂРµСЃС‚Р°Р» С…РѕРґРёС‚СЊ',
        'lost': 'РџРѕС‚РµСЂСЏР»СЃСЏ'
    }
    return status_map.get(status_value, status_value)


@app.get("/api/export/clients")
def export_clients(
    format: str = "csv",
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Export all clients to CSV or XLSX"""
    clients = db.query(Client).order_by(Client.name.asc()).all()
    
    # Get all partners for referral names
    partners = {p.id: p.name for p in db.query(Partner).all()}
    
    headers = ["Р¤РРћ", "РўРµР»РµС„РѕРЅ", "Р”Р°С‚Р° СЂРѕР¶РґРµРЅРёСЏ", "Р’РѕР·СЂР°СЃС‚", "РџРѕР»", "РђРґСЂРµСЃ", 
               "РљР°Рє СѓР·РЅР°Р»Рё", "РЎС‚Р°С‚СѓСЃ", "РџСЂРёС‡РёРЅР° СѓС…РѕРґР°", "Р”Р°С‚Р° СЃРѕР·РґР°РЅРёСЏ"]
    
    rows = []
    for c in clients:
        referral = ""
        if c.referral_partner_id and c.referral_partner_id in partners:
            referral = partners[c.referral_partner_id]
        elif c.referral_custom:
            referral = c.referral_custom
        
        rows.append([
            c.name or "",
            c.phone or "",
            format_date(c.birth_date),
            str(c.age) if c.age else "",
            c.gender or "",
            c.address or "",
            referral,
            get_status_label(c.status),
            c.stopped_reason or "",
            format_datetime(c.created_at)
        ])
    
    if format == "xlsx":
        return generate_xlsx(headers, rows, "clients")
    return generate_csv(headers, rows, "clients")


@app.get("/api/export/partners")
def export_partners(
    format: str = "csv",
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Export all partners to CSV or XLSX"""
    partners = db.query(Partner).order_by(Partner.name.asc()).all()
    
    headers = ["РќР°Р·РІР°РЅРёРµ", "РљРѕРЅС‚Р°РєС‚С‹", "РўРёРї", "РЈСЃР»РѕРІРёСЏ", "РљРѕРјРјРµРЅС‚Р°СЂРёР№", "Р”Р°С‚Р° СЃРѕР·РґР°РЅРёСЏ"]
    
    rows = []
    for p in partners:
        rows.append([
            p.name or "",
            p.contacts or "",
            p.type or "",
            p.terms or "",
            p.comment or "",
            format_datetime(p.created_at)
        ])
    
    if format == "xlsx":
        return generate_xlsx(headers, rows, "partners")
    return generate_csv(headers, rows, "partners")


@app.get("/api/clients/{client_id}/export/sessions")
def export_sessions(
    client_id: int,
    format: str = "csv",
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Export sessions for a client to CSV or XLSX"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    sessions = db.query(LaserSession).filter(
        LaserSession.client_id == client_id
    ).order_by(LaserSession.session_date.asc()).all()
    
    # Get tattoo names
    tattoos = {t.id: t.name for t in db.query(Tattoo).filter(Tattoo.client_id == client_id).all()}
    
    headers = ["РўР°С‚СѓРёСЂРѕРІРєР°/РўР°С‚СѓР°Р¶", "в„– СЃРµР°РЅСЃР°", "РЈС‡Р°СЃС‚РѕРє", "Р”Р»РёРЅР° РІРѕР»РЅС‹", "Р”РёР°РјРµС‚СЂ", 
               "РџР»РѕС‚РЅРѕСЃС‚СЊ", "Р“РµСЂС†", "Р’СЃРїС‹С€РєРё", "Р”Р°С‚Р° СЃРµР°РЅСЃР°", "РџРµСЂРµСЂС‹РІ", "РљРѕРјРјРµРЅС‚Р°СЂРёР№"]
    
    rows = []
    for s in sessions:
        tattoo_name = ""
        if s.tattoo_id and s.tattoo_id in tattoos:
            tattoo_name = tattoos[s.tattoo_id]
        elif s.tattoo_name:
            tattoo_name = s.tattoo_name
        
        rows.append([
            tattoo_name,
            str(s.session_number) if s.session_number else "",
            s.sub_session or "",
            str(s.wavelength) if s.wavelength else "",
            str(s.diameter) if s.diameter else "",
            str(s.density) if s.density else "",
            str(s.hertz) if s.hertz else "",
            str(s.flashes_count) if s.flashes_count else "",
            format_date(s.session_date),
            s.break_period or "",
            s.comment or ""
        ])
    
    filename = f"sessions_{client.name.replace(' ', '_')}"
    if format == "xlsx":
        return generate_xlsx(headers, rows, filename)
    return generate_csv(headers, rows, filename)


@app.get("/api/clients/{client_id}/export/tattoos")
def export_tattoos(
    client_id: int,
    format: str = "csv",
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Export tattoos for a client to CSV or XLSX"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="РљР»РёРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    
    tattoos = db.query(Tattoo).filter(
        Tattoo.client_id == client_id
    ).order_by(Tattoo.name.asc()).all()
    
    headers = ["РќР°Р·РІР°РЅРёРµ", "Р—РѕРЅР° СѓРґР°Р»РµРЅРёСЏ", "РљРѕСЂСЂРµРєС†РёР№", "РџРѕСЃР»РµРґРЅРёР№ РїРёРіРјРµРЅС‚", 
               "РџРѕСЃР»РµРґРЅРёР№ Р»Р°Р·РµСЂ", "РќРµ СѓРґР°Р»СЏР» СЂР°РЅРµРµ", "Р“РґРµ СѓРґР°Р»СЏР»Рё", "Р–РµР»Р°РµРјС‹Р№ СЂРµР·СѓР»СЊС‚Р°С‚", "Р”Р°С‚Р° СЃРѕР·РґР°РЅРёСЏ"]
    
    rows = []
    for t in tattoos:
        rows.append([
            t.name or "",
            t.removal_zone or "",
            t.corrections_count or "",
            format_date(t.last_pigment_date),
            format_date(t.last_laser_date),
            "Р”Р°" if t.no_laser_before else "РќРµС‚",
            t.previous_removal_place or "",
            t.desired_result or "",
            format_datetime(t.created_at)
        ])
    
    filename = f"tattoos_{client.name.replace(' ', '_')}"
    if format == "xlsx":
        return generate_xlsx(headers, rows, filename)
    return generate_csv(headers, rows, filename)


def generate_csv(headers, rows, filename):
    """Generate CSV file response"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(headers)
    writer.writerows(rows)
    output.seek(0)
    
    # Use RFC 5987 encoding for non-ASCII filenames
    encoded_filename = quote(f"{filename}.csv")
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


def generate_xlsx(headers, rows, filename):
    """Generate XLSX file response"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Use RFC 5987 encoding for non-ASCII filenames
    encoded_filename = quote(f"{filename}.xlsx")
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )




# ============== Public Booking Routes ==============


def normalize_phone(phone: str) -> str:
    """Normalize phone to 10 digits (without country prefix)."""
    digits = ''.join(filter(str.isdigit, phone or ''))
    if len(digits) == 11 and digits[0] in ('7', '8'):
        digits = digits[1:]
    if len(digits) != 10:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Phone must contain 10 digits without +7/8 prefix"
        )
    return digits


def clean_required(value: str, max_len: int = 500) -> str:
    """Trim and validate required text field."""
    cleaned = (value or '').strip()
    if not cleaned:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="One of required fields is missing"
        )
    return cleaned[:max_len]


def is_not_filled(value) -> bool:
    """Check whether a field is empty or None."""
    return value is None or not str(value).strip()


def is_negative_answer(value: str) -> bool:
    """Detect 'no' answers in RU/EN form variants."""
    lowered = (value or '').strip().lower()
    return lowered in {'\u043d\u0435\u0442', 'no', 'none', '-'}


@app.post("/api/public/booking", response_model=PublicBookingResponse)
def create_public_booking(booking: PublicBookingCreate, db: Session = Depends(get_db)):
    """Create or update client from landing form and save full intake submission."""
    if booking.website and booking.website.strip():
        raise HTTPException(status_code=400, detail="Invalid request")

    if not booking.consent_personal_data:
        raise HTTPException(status_code=400, detail="Consent is required")

    if booking.form_started_at:
        elapsed_ms = int(time.time() * 1000) - int(booking.form_started_at)
        if elapsed_ms < 2000:
            raise HTTPException(status_code=400, detail="Form submitted too fast")

    full_name = clean_required(booking.full_name, 255)
    normalized_phone = normalize_phone(booking.phone)
    address = clean_required(booking.address, 500)
    referral_source = clean_required(booking.referral_source, 255)
    tattoo_type = clean_required(booking.tattoo_type, 255)
    tattoo_age = clean_required(booking.tattoo_age, 255)
    corrections_info = clean_required(booking.corrections_info, 500)
    previous_removal_info = clean_required(booking.previous_removal_info, 500)
    previous_removal_where = clean_required(booking.previous_removal_where, 500)
    desired_result = clean_required(booking.desired_result, 500)

    client = db.query(Client).filter(Client.phone == normalized_phone).first()
    is_new_client = client is None

    if is_new_client:
        client = Client(
            name=full_name,
            phone=normalized_phone,
            birth_date=booking.birth_date,
            address=address,
            referral_custom=referral_source,
            status=ClientStatus.ACTIVE.value
        )
        db.add(client)
        db.flush()
    else:
        if is_not_filled(client.name):
            client.name = full_name
        if client.birth_date is None and booking.birth_date:
            client.birth_date = booking.birth_date
        if is_not_filled(client.address):
            client.address = address
        if is_not_filled(client.referral_custom):
            client.referral_custom = referral_source

    tattoo = db.query(Tattoo).filter(
        Tattoo.client_id == client.id,
        Tattoo.name == tattoo_type
    ).first()

    if tattoo is None:
        tattoo = Tattoo(
            client_id=client.id,
            name=tattoo_type,
            corrections_count=corrections_info,
            no_laser_before=is_negative_answer(previous_removal_info),
            previous_removal_place=previous_removal_where,
            desired_result=desired_result
        )
        db.add(tattoo)
        db.flush()
    else:
        if is_not_filled(tattoo.corrections_count):
            tattoo.corrections_count = corrections_info
        if is_not_filled(tattoo.previous_removal_place):
            tattoo.previous_removal_place = previous_removal_where
        if is_not_filled(tattoo.desired_result):
            tattoo.desired_result = desired_result

    payload_dict = booking.model_dump(mode='json')
    payload_dict['phone_normalized'] = normalized_phone

    submission = IntakeSubmission(
        client_id=client.id,
        tattoo_id=tattoo.id if tattoo else None,
        full_name=full_name,
        phone=normalized_phone,
        birth_date=booking.birth_date,
        address=address,
        referral_source=referral_source,
        tattoo_type=tattoo_type,
        tattoo_age=tattoo_age,
        corrections_info=corrections_info,
        previous_removal_info=previous_removal_info,
        previous_removal_where=previous_removal_where,
        desired_result=desired_result,
        source='landing',
        is_new_client=is_new_client,
        raw_payload=json.dumps(payload_dict, ensure_ascii=False)
    )
    db.add(submission)

    db.commit()

    return PublicBookingResponse(
        success=True,
        message='Booking request saved',
        client_id=client.id,
        is_new_client=is_new_client
    )


@app.get("/booking", response_class=HTMLResponse)
@app.get("/book", response_class=HTMLResponse)
def booking_page(request: Request):
    """Public landing page with booking form."""
    return templates.TemplateResponse("landing.html", {"request": request})


# ============== Page Routes ==============

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    """Main page - redirects to login if not authenticated"""
    if not get_current_user(request):
        return templates.TemplateResponse("login.html", {"request": request})
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    """Login page"""
    if get_current_user(request):
        return templates.TemplateResponse("index.html", {"request": request})
    return templates.TemplateResponse("login.html", {"request": request})


@app.get("/clients/{client_id}/sessions", response_class=HTMLResponse)
def sessions_page(request: Request, client_id: int):
    """Laser sessions page for a client"""
    if not get_current_user(request):
        return templates.TemplateResponse("login.html", {"request": request})
    return templates.TemplateResponse("sessions.html", {"request": request, "client_id": client_id})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)

